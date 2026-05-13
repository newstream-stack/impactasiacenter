import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────
def load(lang, name):
    with open(f"src/i18n/translations/{lang}/{name}.json", encoding="utf-8") as f:
        return json.load(f)

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
SECTION_FILL = PatternFill("solid", fgColor="2D5986")
ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def make_sheet(title):
    if wb.active and wb.active.title == "Sheet":
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)
    return ws

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_header(ws, cols):
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
    ws.row_dimensions[ws.max_row].height = 22

def write_section(ws, label, span=4):
    ws.append([label])
    ws.merge_cells(start_row=ws.max_row, start_column=1,
                   end_row=ws.max_row, end_column=span)
    cell = ws.cell(ws.max_row, 1)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[ws.max_row].height = 18

def write_row(ws, row_data, alt=False):
    ws.append(row_data)
    fill = ALT_FILL if alt else WHITE_FILL
    r = ws.max_row
    for c in range(1, len(row_data)+1):
        cell = ws.cell(r, c)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER
    ws.row_dimensions[r].height = 60

# ── Sheet 1: All translations ────────────────────────────────────────────────
ws = make_sheet("中英對照總表")
set_col_widths(ws, [28, 55, 55, 18])
write_header(ws, ["Key", "中文", "English", "檔案"])

def row(key, zh, en, file, alt=False):
    write_row(ws, [key, zh, en, file], alt)

# common.json
write_section(ws, "導覽列 / Navigation  (common.json)")
zh = load("zh","common"); en = load("en","common")
keys = ["navVision","navIAAIntro","navPhoenix","navPresidium","navThemes","navVenue",
        "navSupport","btnRegister","btnSupportOnline","countdownLabel",
        "ariaBackToTop","ariaClose","ariaMenu"]
for i,k in enumerate(keys):
    row(k, zh.get(k,""), en.get(k,""), "common.json", i%2==1)

# hero.json
write_section(ws, "Hero 主視覺  (hero.json)")
zh = load("zh","hero")["hero"]; en = load("en","hero")["hero"]
for i,k in enumerate(["subtitle","title","eventName"]):
    row(f"hero.{k}", zh[k], en[k], "hero.json", i%2==1)

# about.json — vision
write_section(ws, "異象 / Vision  (about.json)")
zh = load("zh","about"); en = load("en","about")
for i,k in enumerate(["verse1","verse2","ref"]):
    row(f"vision.{k}", zh["vision"][k], en["vision"][k], "about.json", i%2==1)

# about — summary / btnMore / heading / intro
write_section(ws, "關於大會 / About  (about.json)")
row("about.summary",     zh["about"]["summary"],              en["about"]["summary"],              "about.json", False)
row("about.btnMore",     zh["about"]["btnMore"],              en["about"]["btnMore"],              "about.json", True)
row("about.detail.heading", zh["about"]["detail"]["heading"], en["about"]["detail"]["heading"],    "about.json", False)
row("about.detail.intro",   zh["about"]["detail"]["intro"],   en["about"]["detail"]["intro"],      "about.json", True)
for i,p in enumerate(zh["about"]["detail"]["points"]):
    ep = en["about"]["detail"]["points"][i]
    row(f"about.detail.points[{i}].title", p["title"], ep["title"], "about.json", i%2==0)
    row(f"about.detail.points[{i}].desc",  p["desc"],  ep["desc"],  "about.json", i%2==1)

# iaaIntro.json
write_section(ws, "IAA 介紹  (iaaIntro.json)")
zh = load("zh","iaaIntro"); en = load("en","iaaIntro")
row("trailer.title",   zh["trailer"]["title"],    en["trailer"]["title"],    "iaaIntro.json", False)
row("iaaIntro.title",  zh["iaaIntro"]["title"],   en["iaaIntro"]["title"],   "iaaIntro.json", True)
for i,b in enumerate(zh["iaaIntro"]["blocks"]):
    eb = en["iaaIntro"]["blocks"][i]
    row(f"iaaIntro.blocks[{i}].title", b["title"], eb["title"], "iaaIntro.json", i%2==0)
    for j,(p,ep) in enumerate(zip(b["paragraphs"], eb["paragraphs"])):
        row(f"iaaIntro.blocks[{i}].paragraphs[{j}]", p, ep, "iaaIntro.json", j%2==1)

# themes.json
write_section(ws, "大會議題 / Themes  (themes.json)")
zh = load("zh","themes"); en = load("en","themes")
row("themesSection.title", zh["themesSection"]["title"], en["themesSection"]["title"], "themes.json", False)
for i,t in enumerate(zh["themes"]):
    et = en["themes"][i]
    row(f"themes[{i}].title",         t["title"],            et["title"],            "themes.json", True)
    row(f"themes[{i}].summary",       t["summary"],          et["summary"],          "themes.json", False)
    row(f"themes[{i}].detail.heading",t["detail"]["heading"],et["detail"]["heading"],"themes.json", True)
    row(f"themes[{i}].detail.intro",  t["detail"]["intro"],  et["detail"]["intro"],  "themes.json", False)
    for j,p in enumerate(t["detail"]["points"]):
        ep = et["detail"]["points"][j]
        row(f"themes[{i}].points[{j}].title", p["title"], ep["title"], "themes.json", j%2==1)
        row(f"themes[{i}].points[{j}].desc",  p["desc"],  ep["desc"],  "themes.json", j%2==0)

# timeline.json
write_section(ws, "歷屆年會 / Timeline  (timeline.json)")
zh = load("zh","timeline"); en = load("en","timeline")
row("timelineSection.title", zh["timelineSection"]["title"], en["timelineSection"]["title"], "timeline.json", False)
for i,t in enumerate(zh["timeline"]):
    et = en["timeline"][i]
    row(f"timeline[{i}].region",      t.get("region",""),      et.get("region",""),      "timeline.json", i%2==1)
    row(f"timeline[{i}].description", t.get("description",""), et.get("description",""), "timeline.json", i%2==0)
    if "detail" in t:
        d = t["detail"]; ed = et["detail"]
        row(f"timeline[{i}].detail.edition", d["edition"], ed["edition"], "timeline.json", True)
        row(f"timeline[{i}].detail.theme",   d["theme"],   ed["theme"],   "timeline.json", False)
        row(f"timeline[{i}].detail.tagline", d["tagline"], ed["tagline"], "timeline.json", True)
        for j,(p,ep) in enumerate(zip(d["paragraphs"], ed["paragraphs"])):
            row(f"timeline[{i}].detail.paragraphs[{j}]", p, ep, "timeline.json", j%2==0)

# phoenix.json
write_section(ws, "鳳凰城 / Phoenix  (phoenix.json)")
zh = load("zh","phoenix")["phoenix"]; en = load("en","phoenix")["phoenix"]
row("phoenix.subtitle", zh["subtitle"], en["subtitle"], "phoenix.json", False)
row("phoenix.title",    zh["title"],    en["title"],    "phoenix.json", True)
for i,(s,es) in enumerate(zip(zh["stats"], en["stats"])):
    row(f"phoenix.stats[{i}]", f"{s['value']}  {s['label']}", f"{es['value']}  {es['label']}", "phoenix.json", i%2==0)
for i,(intro,ei) in enumerate(zip(zh["intro"], en["intro"])):
    row(f"phoenix.intro[{i}].label", intro["label"], ei["label"], "phoenix.json", i%2==1)
    row(f"phoenix.intro[{i}].desc",  intro["desc"],  ei["desc"],  "phoenix.json", i%2==0)
for i,cat in enumerate(zh["categories"]):
    ecat = en["categories"][i]
    row(f"phoenix.categories[{i}].tag",     cat["tag"],     ecat["tag"],     "phoenix.json", True)
    row(f"phoenix.categories[{i}].heading", cat["heading"], ecat["heading"], "phoenix.json", False)
    for j,item in enumerate(cat["items"]):
        eitem = ecat["items"][j]
        row(f"phoenix.categories[{i}].items[{j}].label", item["label"], eitem["label"], "phoenix.json", j%2==1)
        row(f"phoenix.categories[{i}].items[{j}].desc",  item["desc"],  eitem["desc"],  "phoenix.json", j%2==0)
row("phoenix.outro", zh["outro"], en["outro"], "phoenix.json", False)

# presidium.json
write_section(ws, "主席團 / Presidium  (presidium.json)")
zh = load("zh","presidium")["presidium"]; en = load("en","presidium")["presidium"]
row("presidium.title",        zh["title"],        en["title"],        "presidium.json", False)
row("presidium.subtitle",     zh["subtitle"],     en["subtitle"],     "presidium.json", True)
row("presidium.coChairsTitle",zh["coChairsTitle"],en["coChairsTitle"],"presidium.json", False)
for i,m in enumerate(zh["main"]):
    em = en["main"][i]
    row(f"presidium.main[{i}].role",  m["role"],            em["role"],  "presidium.json", i%2==1)
    row(f"presidium.main[{i}].name",  m["name"],            em["name"],  "presidium.json", i%2==0)
    row(f"presidium.main[{i}].title", m["title"],           em["title"], "presidium.json", i%2==1)
    row(f"presidium.main[{i}].bio",   m.get("bio",""),      em.get("bio",""), "presidium.json", i%2==0)
for i,m in enumerate(zh["coChairs"]):
    em = en["coChairs"][i]
    row(f"presidium.coChairs[{i}].name",  m["name"],          em["name"],          "presidium.json", i%2==1)
    row(f"presidium.coChairs[{i}].title", m["title"],         em["title"],         "presidium.json", i%2==0)
    row(f"presidium.coChairs[{i}].bio",   m.get("bio",""),    em.get("bio",""),    "presidium.json", i%2==1)

# venue.json
write_section(ws, "會場 / Venue  (venue.json)")
zh = load("zh","venue")["venue"]; en = load("en","venue")["venue"]
for i,k in enumerate(["title","name","heading","desc","date"]):
    row(f"venue.{k}", zh[k], en[k], "venue.json", i%2==1)

# footer.json
write_section(ws, "頁尾 / Footer  (footer.json)")
zh = load("zh","footer")["footer"]; en = load("en","footer")["footer"]
for i,k in enumerate(["brandDesc","quickLinks","contact","email","phone","copyright"]):
    row(f"footer.{k}", zh[k], en[k], "footer.json", i%2==1)

# freeze top row
ws.freeze_panes = "A2"

# ── Save ─────────────────────────────────────────────────────────────────────
out = "translations_比對表.xlsx"
wb.save(out)
print(f"✅ 已產生：{out}")
